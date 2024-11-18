#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
模块名称：doubao_parallel_test
模块描述：提供基于多线程对doubao模型进行大规模数据测试的能力

作者：Albert Liu
更新日期：2024-11-18
版本：v5.0

update log
    - v5.0
        1. sp和up增加字符串泛型约束
        2. 解决并发控制失效问题
        3. 修改时间记录逻辑
        4. 优化日志处理逻辑
        5. 增加多轮对话测试能力
        6. 修改配置文件设定
    - v4.0
        1. 使用python coroutine重构代码，提升代码运行效率
        2. 大幅缩减为可执行任务种类为2个，增强代码可读性便于自定义修改逻辑
        3. 增加简易demo，上手即用
        4. 更新Ark sdk版本，升级到v3，兼容openai sdk
        5. bug修复
        - 解决前序版本数据集读取重复导致内存占用过大等问题
        - 解决v2.0版本加锁异常
    - v3.0
        1. 修复并发启动时出现部分thread启动失败问题
        2. 修改并发逻辑，处理单文件并发写入问题
    - v2.0 增加多线程并发测试能力
    - v1.0 实现基本逻辑

依赖：
    - openpyxl (安装方法: `pip3 install openpyxl`)
    - volcengine.maas.v3 (安装方法: `pip3 install volcengine`)

用法示例：
    使用方式：
    1. 低代码使用方法：
        1.0 从以下两种选择中，选择测试内容，修改config.ini文件中的test_data_name文件名
         - 单轮对话测试: multi_round.xlsx
         - 多轮对话测试: single_round.xlsx
        1.1 替换文件中数据待测试数据
            - 单轮对话：填充后的文档包含3个字段，如[test_index（从0开始）, system_prompt, user_prompt_1]
            - 多轮对话：填充后的文档每行包含一次与模型的完整多轮对话，每轮对话占一个单元格，如[test_index（从0开始）, system_prompt, up1,ap1,up2,ap2,up3,ap3....upX]，注意最后一个一定是upX
                     （注：多轮对话的对话轮次允许不一致，如第一行只有3轮对话，第二行允许只有1轮对话）
        1.2 修改config.ini中api_key及endpoint_id
        1.3 选择以下两种方式其一，运行代码
         - order_test: 顺序查询，从0开始，查询到CONSTANTS.total_test_data_num，每个case查询CONSTANTS.try_times次
         - index_test: 点查，按照test_index_list中要求的index进行查询，每个case查询CONSTANTS.try_times次
        1.4 结果生成在./result/中
         - 单轮对话：single_round_v0.xlsx
         - 多轮对话：multi_round_v0.xlsx
    2. 高代码使用方法：
        2.1 支持以下4个方法重构，重构建议在方法注释中
            - load_test_data: 自定义测试集数据读取逻辑，适应不同的测试集数据格式
            - read_system_prompt: 自定义sp的组装格式，适应变量替换 or sp不一致场景
            - resolve_user_ass_prompt: 自定义up及ap的解析格式，适应up读取数据后，需要二次增加内容
            - result_process: 自定义结果存储格式，适应于格式化输出如json需解析内容
        2.2 如需要增加全局参数，考虑在config.ini定义后，在Constant class中读取
    3. 结果输出
    结果都生成在work_directory/result_directory路径下，一般会生成3类文件
    - data.xlsx: 存储模型运算结果数据 及 result_process后处理后的自定义结果数据
    - log.log: 运行日志
    - wrong.txt: 存储多次重试后，仍失败的任务index（test_data中的index）

    ☆注意：
    1. 建议只修改以上4个推荐修改的函数，其他函数涉及到coroutine调度逻辑，建议不要动
    2. 核心代码逻辑在request函数中
"""
import asyncio
import logging
import traceback
import os
import json
import configparser
import openpyxl
import time
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from typing import Optional, Any
from volcenginesdkarkruntime import AsyncArk

from tts.tts_http_demo_origin import header

"""========================================================================================="""
"""==                        以下函数需要根据测试集情况，自定义解析逻辑                           =="""
"""========================================================================================="""


# 读取测试机
def load_test_data(file_path: str) -> list:
    """
    读取测试集，当测试集为json list文件时，可自定义加载逻辑

    :param file_path:
    :return: 以list形式组装test_data数据集，具体test_data数据集的解析结构在read_user_prompt方法中定义
    """
    ## 读取excel中的数据
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook['Sheet1']
    worksheet._current_row = worksheet.max_row
    # 初始化二维列表
    data = []
    # 遍历工作表中的所有行和列，并将数据添加到二维列表中
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


# 读取system_prompt
def read_system_prompt(file_path: str, test_index: int) -> str:
    """
    构建system_prompt，有以下两种方式
    1. 从测试数据集中根据test_index取
    2. 从system_prompt.txt文件中读取
    默认读取当前目录下的system_prompt.txt文件。
    注意：
        如果需要对system_prompt做自定义变量替换，在这个函数中做修改即可，推荐使用以下方式替换变量
        "xxx{param1}xxx{param2}xxx".format(param1 = A,param2 = B)

    :param file_path:
    :param test_index: 测试数据集中的本轮测试用例的index，用于去Test_Data_List中取数据
    :return system_prompt:
    """
    # 读取excel第一列
    item = CONSTANTS.test_data_list[test_index]
    return item[1]

    # 读取file_path下的system_prompt.txt文件中
    # try:
    #     with open(file_path, 'r', encoding='utf-8') as file:
    #         system_prompt = file.read()
    #     return system_prompt
    # except FileNotFoundError:
    #     CONSTANTS.logger.error("system_prompt文件未找到")


# 从测试集中解析出每一轮的user_prompt
def resolve_user_ass_prompt(test_index: int) -> list[str]:
    """
    从read_test_data得到的测试数据集Test_Data_List中解析出user_prompt，逻辑需要自定义

    :param test_index: 测试数据集中的本轮测试用例的index，用于去Test_Data_List中取数据
    :return: 返回一个list
            - 单轮对话：只含有1个元素的list，格式如[up1]
            - 多轮对话：组装完成后的多轮up和ap组成的list，格式如[up1,ap1,up2,ap2,up3,ap3....upX]，注意最后一个一定是upX
    """
    item = CONSTANTS.test_data_list[test_index]
    return item[2:]


# doubao请求后处理
def result_process(user_ass_prompt: list[str], raw_resp: str, test_index: int) -> list[str]:
    """
    对doubao调用结果进行处理，处理后的内容也会写入excel中。默认不做任何修改写入excel

    :param user_ass_prompt: 历史prompt list
    :param raw_resp: 单次调用doubao接口的直接返回结果str，待处理内容
    :param test_index: 测试数据集中的本轮测试用例的index，用于去Test_Data_List中取数据
    :return: 处理后的结果list[str]
    """
    # proc_result = [raw_resp]
    proc_result = []

    return proc_result


"""========================================================================================="""
"""==                             以下逻辑非特殊情况，完全不需要动                              =="""
"""========================================================================================="""

"""========================================= 写入处理 ======================================="""


# 顺序测试
def order_test():
    """
    顺序查询，从0开始，查询到CONSTANTS.total_test_data_num，每个case查询CONSTANTS.try_times次

    :return: test结果
    """
    try:
        # test数据集的index列表
        test_index_list = [i for i in range(0, CONSTANTS.total_test_data_num)]
        # 多并发运行
        results = asyncio.run(run_tasks(test_index_list))
    finally:
        # 记录wrong index
        wrong_num = write_wrong(CONSTANTS.wrong_queue)
    return results


# 序号点查
def index_test(test_index_list: list[int]):
    """
    点查，按照test_index_list中要求的index进行查询，每个case查询CONSTANTS.try_times次

    :param test_index_list:
    :return:
    """
    try:
        # 多并发运行
        results = asyncio.run(run_tasks(test_index_list))
    finally:
        # 记录wrong index
        wrong_num = write_wrong(CONSTANTS.wrong_queue)
    return results


"""========================================= 请求处理 ======================================="""


# request构建
async def request(excel_index: int, test_index: int) -> Optional[str]:
    """
    包装doubao方法，构成单次coroutine调用

    :param excel_index: excel中的index
    :param test_index:  测试数据集index
    :return: 单次调用doubao接口的返回结果str
    """
    # 记录并发数量
    CONSTANTS.current_running_tasks += 1
    # # 记录开始时间
    # start_time = time.time()
    # CONSTANTS.logger.info(
    #     f"START: task_id(excel_index): {excel_index} | test_index: {test_index} %%%% | start_time: {start_time} | concurrent_running_task: {CONSTANTS.current_running_tasks}")
    for attempt_round in range(CONSTANTS.retry_times):
        raw_resp = None
        try:
            # 构建sp和up
            system_prompt = str(read_system_prompt(CONSTANTS.system_prompt_name, test_index))
            user_ass_prompt = resolve_user_ass_prompt(test_index)
            # 请求doubao
            raw_resp = await doubao(system_prompt, user_ass_prompt)
            # 结果后处理
            processed_resp = result_process(user_ass_prompt, raw_resp, test_index)
            # 组装结果 [serial_number, repeat_time, system_prompt, user_ass_prompt， assistant_answer] + 自定义数据
            excel_content = [excel_index, test_index, system_prompt] + user_ass_prompt + [raw_resp] + processed_resp
            # 写入excel
            async with CONSTANTS.excel_lock:
                write_excel(excel_content)
            # 记录时间
            end_time = time.time()
            # 记录并发数量
            CONSTANTS.current_running_tasks -= 1
            # 打印日志
            CONSTANTS.logger.info(
                f"SUCCESS: task_id(excel_index): {excel_index} | test_index: {test_index} | attempt_round: {attempt_round} %%%% total_time: {end_time-CONSTANTS.zero_time} | concurrent_running_task: {CONSTANTS.current_running_tasks}")
            return raw_resp
        except Exception as e:
            # 记录时间
            end_time = time.time()
            if attempt_round < CONSTANTS.retry_times - 1:
                # 记录错误日志
                CONSTANTS.logger.error(
                    f"FAILED: task_id(excel_index): {excel_index} | test_index: {test_index} | attempt_round: {attempt_round} %%%% total_time: {end_time-CONSTANTS.zero_time} | concurrent_running_task: {CONSTANTS.current_running_tasks}")
                # 间隔一定时长重试
                await asyncio.sleep(CONSTANTS.retry_delay)
            else:
                # 记录并发数量
                CONSTANTS.current_running_tasks -= 1
                # 记录错误日志
                CONSTANTS.logger.error(
                    (
                        f"FINAL_FAILED: task_id(excel_index): {excel_index} | test_index: {test_index} | attempt_round: {attempt_round} %%%% total_time: {end_time-CONSTANTS.zero_time} | concurrent_running_task: {CONSTANTS.current_running_tasks}\n"
                        f"======================== ExceptionMessage ========================: \n"
                        f"{str(e)}\n"
                        f"======================== TraceBack ========================: \n"
                        f"{traceback.format_exc()}"
                        f"======================== RawCntent ========================: \n"
                        f"{raw_resp}"
                    )
                )
                # 记录错误index，用于重新跑数据
                await CONSTANTS.wrong_queue.put(test_index)
    return None


# doubao调用
async def doubao(system_prompt: str, user_ass_prompt: list[str]) -> str:
    """
    doubao异步请求

    :param system_prompt:
    :param user_ass_prompt:
    :return: 单次调用doubao接口的返回结果str
    """
    user_ass_prompt = [s for s in user_ass_prompt if s is not None]
    messages = [
        {"role": "system", "content": system_prompt}
    ]
    for i, item in enumerate(user_ass_prompt):
        if i % 2 == 0:
            messages.append(
                {"role": "user", "content": item}
            )
        elif i % 2 == 1:
            messages.append(
                {"role": "assistant", "content": item}
            )

    completion = await CONSTANTS.client.chat.completions.create(
        model=CONSTANTS.endpoint_id,
        messages=messages,
        stream=False,
        logprobs=False,
        max_tokens=CONSTANTS.max_tokens,
        frequency_penalty=CONSTANTS.frequency_penalty,
        presence_penalty=CONSTANTS.presence_penalty,
        temperature=CONSTANTS.temperature,
        top_p=CONSTANTS.top_p,
        top_logprobs=CONSTANTS.top_logprobs,
        n=1,
    )
    return completion.choices[0].message.content


# 整体coroutine调度
async def run_tasks(test_index_list: list[int]):
    """
    coroutine协调函数，控制循环

    :param test_index_list: 待测试的test_index
    :return: 所有调用doubao接口的返回结果list[str]
    """
    # 并发控制
    semaphore = asyncio.Semaphore(CONSTANTS.thread_num)
    # 每个index复制CONSTANTS.try_times份（测试CONSTANTS.try_times次）
    total_index_list = [element for element in test_index_list for _ in range(CONSTANTS.try_times)]
    # 执行并行运算任务
    tasks = [concurrent_control(semaphore, excel_index, test_index) for excel_index, test_index in
             enumerate(total_index_list)]
    return await asyncio.gather(*tasks)


# 并发控制
async def concurrent_control(semaphore: asyncio.Semaphore, excel_index: int, test_index: int) -> Optional[str]:
    """
    此函数用于coroutine并发数限制

    :param concurrency_limit: 并发数限制
    :param excel_index: excel中的index，同时也是task_id
    :param test_index:  测试数据集index
    :return: 单次调用doubao接口的返回结果str
    """
    async with semaphore:
        return await request(excel_index, test_index)


"""========================================= 写入处理 ======================================="""


# 结果写入excel
def write_excel(row_list: list):
    """
    结果写入excel
    :param row_list: 一行结果内容
    :return: None
    """
    # 从第二行开始写
    for cindex, item in enumerate(row_list):
        CONSTANTS.excel_info["worksheet"].cell(row=row_list[0] + 2, column=cindex + 1, value=item)
    # 保存结果
    CONSTANTS.excel_info["workbook"].save(CONSTANTS.excel_info["excel_path"])


# wrong_index结果写入文档
def write_wrong(queue: asyncio.Queue) -> list[str]:
    """
    对于超过CONSTANTS.retry_times仍计算错误的test_index，会记录到CONSTANTS.result_directory/CONSTANTS.wrong_name文件中

    :param queue: wrong_index的记录队列
    :return: 记录wrong_index的list
    """
    # 获取wrong index
    wrong_items = []
    seen = set()  # 用于存储已经遇到的元素
    while not queue.empty():
        item = queue.get_nowait()
        if item is None:  # 结束标志
            break
        if item not in seen:
            seen.add(item)
            wrong_items.append(str(item))
    # 非空创建
    if len(wrong_items) != 0:
        # 排序
        wrong_items.sort()
        # 写入文件
        with open(os.path.join(CONSTANTS.result_directory, CONSTANTS.wrong_name), "w") as num_file:
            num_file.write(f"[{",".join(wrong_items)}]")
        return wrong_items


"""========================================= 启动参数处理 ======================================="""


# 定义参数管理类
class Constants:
    # 配置文件数据
    ## 存储路径
    work_directory = None
    test_directory = None
    system_prompt_name = None
    test_data_name = None
    result_directory = None
    result_flag = None
    log_name = None
    wrong_name = None
    ## 模型参数
    api_key = None
    endpoint_id = None
    ## 次数控制
    total_test_data_num = None
    try_times = None
    thread_num = None
    ## 重试
    retry_times = None
    retry_delay = None
    ## 模型参数
    max_tokens = None
    frequency_penalty = None
    presence_penalty = None
    temperature = None
    top_p = None
    top_logprobs = None

    # 其他数据
    ## 读取测试数据集
    test_data_list = None
    ## Excel文件句柄
    excel_info = None
    ## 创建并行锁
    excel_lock = None
    ## 启动日志记录
    logger = None
    ## 记录wrong index的queue
    wrong_queue = None
    ## ark client
    client = None

    # 单例类
    _instance = None

    # 创建单例类
    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            # 如果 _instance 为空，则创建一个新的实例
            cls._instance = super(Constants, cls).__new__(cls)
        return cls._instance

    def __init__(self):
        # 获取配置文件数据
        self.__get_param()
        # 启动其他数据定义
        self.__prepare()
        # 日志组件启动加载
        self.__start_log()
        # 初始化Ark client
        self.client = AsyncArk(api_key=self.api_key)

        if self.total_test_data_num == 0:
            self.total_test_data_num = len(self.test_data_list)

    # 读取配置文件参数
    def __get_param(self):
        """
        配置文件参数加载

        :return: None
        """
        try:
            # 创建配置解析器对象
            config = configparser.ConfigParser()
            # 读取配置文件并指定编码为 utf-8
            with open('config.ini', 'r', encoding='utf-8') as config_file:
                config.read_file(config_file)
            # 常量参数读取
            ## 存储路径
            section = "PATH"
            self.work_directory = config[section].get("work_directory")
            section = "TEST_PATH"
            self.test_directory = config[section].get("test_directory")
            self.system_prompt_name = config[section].get("system_prompt_name")
            self.test_data_name = config[section].get("test_data_name")
            section = "RESULT_PATH"
            self.result_directory = config[section].get("result_directory")
            self.result_flag = config[section].get("result_flag")
            self.log_name = config[section].get("log_name")
            self.wrong_name = config[section].get("wrong_name")
            ## 模型参数
            section = "PARAM"
            self.api_key = config[section].get("api_key")
            self.endpoint_id = config[section].get("endpoint_id")
            ## 次数控制
            section = "NUM"
            self.total_test_data_num = int(config[section].get("total_test_data_num"))
            self.try_times = int(config[section].get("try_times"))
            self.thread_num = int(config[section].get("thread_num"))
            ## 重试
            section = "RETRY"
            self.retry_times = int(config[section].get("retry_times"))
            self.retry_delay = int(config[section].get("retry_delay"))
            ## 模型参数
            section = "MODEL"
            self.max_tokens = int(config[section].get("max_tokens"))
            self.frequency_penalty = float(config[section].get("frequency_penalty"))
            self.presence_penalty = float(config[section].get("presence_penalty"))
            self.temperature = float(config[section].get("temperature"))
            self.top_p = float(config[section].get("top_p"))
            self.top_logprobs = int(config[section].get("top_logprobs"))

            # ## 本次测试数量（若测试数量小于测试集规模，则只取前Test_Data_Num个）
            # if self.total_test_data_num * self.try_times < self.thread_num:
            #     print("测试量级小于Thread_Num")
        except configparser.NoSectionError as e:
            print(f"Error: Get Parameter Wrong, Section not found: {e.section}")
        except configparser.NoOptionError as e:
            print(f"Error: Get Parameter Wrong, Option not found: {e.option}")
        except ValueError as e:
            print(f"Error: Get Parameter Wrong, Value error: {e}")
        except Exception as e:
            print(f"Get Parameter Wrong, Unexpected error: {e}")

    # 启动前数据加载
    def __prepare(self):
        """
        加载完数据后，启动前其他参数处理

        :return:None
        """
        # 路径处理
        ## 刷新输入输出核心工作路径
        self.result_directory = os.path.join(self.work_directory, self.result_directory)
        self.test_directory = os.path.join(self.work_directory, self.test_directory)
        # 路径创建
        if not os.path.exists(self.result_directory):
            os.makedirs(self.result_directory)
        if not os.path.exists(self.test_directory):
            os.makedirs(self.test_directory)
        ## 输入输出完整路径
        test_data_path = os.path.join(self.test_directory, self.test_data_name)
        result_name = self.test_data_name.replace(".xlsx", self.result_flag+".xlsx")
        excel_path = os.path.join(self.result_directory, result_name)

        # 读取测试数据集
        test_data_list_with_title = load_test_data(test_data_path)
        self.test_data_list = test_data_list_with_title[1:]
        self.excel_info = self.__construct_excel(excel_path, test_data_list_with_title[0])
        # 创建并行锁
        self.excel_lock = asyncio.Lock()
        # 启动日志记录
        self.logger = self.__start_log()
        # 创建记录wrong index的queue
        self.wrong_queue = asyncio.Queue()
        # 创建并发记录
        self.current_running_tasks = 0
        # 创建初始时间记录
        self.zero_time = time.time()

    # 日志组件启动加载
    def __start_log(self) -> logging.Logger:
        """
        日志组件启动加载

        :return: 日志组件句柄
        """
        # 配置日志记录器
        logging.basicConfig(
            level=logging.INFO,  # 设置日志级别
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',  # 设置日志格式
            handlers=[
                logging.FileHandler(os.path.join(self.result_directory, self.log_name), encoding='utf-8'),  # 将日志写入文件
                logging.StreamHandler()  # 将日志输出到控制台
            ]
        )
        # 废除其他日志
        logger = logging.getLogger('httpx')
        logger.addHandler(logging.NullHandler())
        # 启动日志记录
        logger = logging.getLogger(f"__main__")
        # 废除多余日志
        logging.getLogger("httpx").setLevel(logging.WARNING)
        self.logger = logger
        return logger

    # 结果存储excel组装
    def __construct_excel(self, excel_path: str, headers: list[str]) -> dict[str, str]:
        """
        构建存放测试结果的excel，指定目录和拿到句柄，实际数据处理在write_excel方法中

        :param excel_path: 结果excel文件目录
        :return excel_info: 当前excel的写入句柄等信息
        """
        sheet_name = "data"
        # 表格处理
        if os.path.exists(excel_path):
            # 数据追加
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook[sheet_name]
            worksheet._current_row = worksheet.max_row
        else:
            # 新创sheet及file
            workbook = Workbook()
            worksheet = workbook.create_sheet(sheet_name, 0)
            # 表头设置
            new_header = ["excel_index（从0开始）"] + headers + ["AP"]
            worksheet.append(new_header)
        # 结果构成
        excel_info = {
            "excel_path": excel_path,
            "worksheet": worksheet,
            "workbook": workbook,
        }
        return excel_info


if __name__ == '__main__':
    # 获取启动准备参数
    CONSTANTS = Constants()

    CONSTANTS.logger.info("开始计算")

    # 以下两种方式，只能选一种使用
    """ 方式一：按照配置文件顺序查询, 从0开始 """
    order_test()
    """ 方式二：按照list点查(test_data中，index从0开始表示第一个例子) """
    # test_index_list = [0, 1, 8, 9]
    # index_tes(test_index_list)

    CONSTANTS.logger.info("计算结束！")
    # print(results)
